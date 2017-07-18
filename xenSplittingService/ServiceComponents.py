# -*- encoding: UTF-8 -*-
# --------------------------------------------------------
import csv
import pandas as pd
from openpyxl import load_workbook
# --------------------------------------------------------


# --------------------------------------------------------
# read and write csv file
def load_csv(filename):
    """
    load a list which is of two dimension
    with lines in list and columns in sub_lists
    """
    csvfile = open(filename, 'r', encoding='utf-8', newline='')
    spamreader = csv.reader(csvfile,
                            delimiter=',',
                            quotechar='"'
                            )
    return [var for var in spamreader]


def save_csv(filename, content):
    """
    save a list which is of two dimension to the file
    with lines in list and columns in sub_lists
    """
    csvfile = open(filename, 'w', encoding='utf-8', newline='')
    spamwriter = csv.writer(csvfile,
                            delimiter=',',
                            quotechar='"',
                            quoting=csv.QUOTE_MINIMAL
                            )
    spamwriter.writerows(content)
    csvfile.close()
    return True
# --------------------------------------------------------


# --------------------------------------------------------
def load_excel(path, sheetname=0, header=0, skiprows=None, skip_footer=0, index_col=None, names=None,
               parse_cols=None, parse_dates=False, date_parser=None, na_values=None, thousands=None,
               convert_float=True, has_index_names=None, converters=None, dtype=None, true_values=None,
               false_values=None, engine=None, squeeze=False):
    """Use pandas.read_excel to load excel files"""
    return pd.read_excel(path, sheetname=sheetname, header=header, skiprows=skiprows, skip_footer=skip_footer,
                         index_col=index_col, names=names, parse_cols=parse_cols, parse_dates=parse_dates,
                         date_parser=date_parser, na_values=na_values, thousands=thousands, convert_float=convert_float,
                         has_index_names=has_index_names, converters=converters, dtype=dtype, true_values=true_values,
                         false_values=false_values, engine=engine, squeeze=squeeze)


class ExcelFileWriter(object):
    def __init__(self, path: str, engine='openpyxl'):
        self.writer = pd.ExcelWriter(path=path, engine=engine)

    def append_sheet(self, content: pd.DataFrame, sheet_name='Sheet1', na_rep=str(), float_format=None,
                     columns=None, header=True, index=True, index_label=None, startrow=0, startcol=0, merge_cells=True,
                     encoding=None, inf_rep='inf', verbose=True, freeze_panes=None):
        content.to_excel(excel_writer=self.writer, sheet_name=sheet_name, na_rep=na_rep, float_format=float_format,
                         columns=columns, header=header, index=index, index_label=index_label, startrow=startrow,
                         startcol=startcol, merge_cells=merge_cells, encoding=encoding, inf_rep=inf_rep,
                         verbose=verbose, freeze_panes=freeze_panes)

    def save(self):
        self.writer.save()

    def finish(self):
        self.writer.save()
        self.writer.close()
        del self
# --------------------------------------------------------


# --------------------------------------------------------
class MultiList(object):
    def __init__(self, components: list):
        self.lists = dict()
        for tag in components:
            if type(tag) != str:
                raise ValueError('MultiList.__init__(components): expected str in parameter components')
            self.lists[tag] = list()

    def contain(self, element):
        flag = False
        for tag in self.lists:
            if element in self.lists[tag]:
                flag = True
        return flag

    def tags(self):
        return self.lists.keys()

    def append(self, tag: str, element):
        try:
            self.lists[tag].append(element)
        except KeyError:
            raise KeyError('MultiList.append(tag, x): tag not in MultiList')

    def remove(self, tag: str, element):
        try:
            self.lists[tag].remove(element)
        except KeyError:
            raise KeyError('MultiList.remove(tag, x): tag not in MultiList')
        except ValueError:
            raise ValueError('MultiList.remove(tag, x): x not in MultiList')
# --------------------------------------------------------


# --------------------------------------------------------
class MultiTable(object):
    def __init__(self, file_path: str, indexing=True):
        excel_book = load_workbook(filename=file_path, read_only=True)
        sheet_name_list = excel_book.get_sheet_names
        self.data_dicts = dict()
        self.data_list_for_search = dict()
        for sheet in sheet_name_list:
            self.data_dicts[sheet] = pd.read_excel(file_path, sheetname=sheet)
        if indexing is True:
            self.__indexing__()

    def __indexing__(self):
        for sheet in self.data_dicts:
            columes = self.data_dicts[sheet].columns.tolist()
            for col in columes:
                values = self.data_dicts[sheet][col].tolist()
                for val in values:
                    self.data_list_for_search[val] = 0

    def contain(self, element):
        if len(self.data_list_for_search) != 0:
            return element in self.data_list_for_search
        else:
            for sheet in self.data_dicts:
                columes = self.data_dicts[sheet].columns.tolist()
                for col in columes:
                    values = self.data_dicts[sheet][col].tolist()
                    if element in values:
                        return True
            return False

    def add(self, sheet: str, colume: str, element: str):
        if sheet in self.data_dicts:
            pass
        else:
            raise ValueError('sheet {0} not in Multitable'.format(sheet))
# --------------------------------------------------------


# --------------------------------------------------------
if __name__ == '__main__':
    pass
